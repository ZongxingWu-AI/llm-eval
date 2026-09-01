"""交互式 JSONL 转 Excel 工具。

每次运行时手动输入 JSONL 输入路径和 XLSX 输出路径。复杂字段保留为
中文 JSON 字符串，便于查看原始结构，同时避免把表格展开得过宽。
"""

from __future__ import annotations

import json
from json import JSONDecodeError
from pathlib import Path
from typing import Any

EXCEL_MAX_CELL_LENGTH = 32_767
_TRUNCATION_SUFFIX = "\n...[内容超过 Excel 单元格长度限制，已截断]"


def _clean_path(value: str) -> Path:
    """清理用户从资源管理器复制的路径，并转换为 Path。"""

    cleaned = value.strip()
    if len(cleaned) >= 2 and cleaned[0] == cleaned[-1] and cleaned[0] in {'"', "'"}:
        cleaned = cleaned[1:-1].strip()
    if not cleaned:
        raise ValueError("路径不能为空")
    return Path(cleaned).expanduser()


def _json_cell_value(value: Any) -> Any:
    """把嵌套 JSON 值转换为可写入 Excel 单元格的内容。"""

    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _truncate_cell_value(value: Any) -> tuple[Any, bool]:
    """将超过 Excel 单元格限制的文本截断，并返回是否发生截断。"""

    if not isinstance(value, str) or len(value) <= EXCEL_MAX_CELL_LENGTH:
        return value, False

    keep_length = EXCEL_MAX_CELL_LENGTH - len(_TRUNCATION_SUFFIX)
    return value[:keep_length] + _TRUNCATION_SUFFIX, True


def _read_jsonl(input_path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    """读取 JSONL，并返回记录与首次出现的字段顺序。"""

    rows: list[dict[str, Any]] = []
    columns: list[str] = []
    try:
        lines = input_path.read_text(encoding="utf-8-sig").splitlines()
    except UnicodeDecodeError as exc:
        raise ValueError("输入文件不是有效的 UTF-8 文本") from exc

    for line_number, line in enumerate(lines, start=1):
        if not line.strip():
            continue
        try:
            row = json.loads(line)
        except JSONDecodeError as exc:
            raise ValueError(f"第 {line_number} 行 JSON 解析失败：{exc.msg}") from exc
        if not isinstance(row, dict):
            raise ValueError(f"第 {line_number} 行不是 JSON 对象，无法作为表格记录")
        rows.append(row)
        for key in row:
            if key not in columns:
                columns.append(key)

    if not rows:
        raise ValueError("输入 JSONL 没有有效记录")
    return rows, columns


def _sheet_title(input_path: Path) -> str:
    """生成符合 Excel 31 字符限制且不含非法字符的工作表名称。"""

    title = input_path.stem or "jsonl"
    for character in "[]:*?/\\":
        title = title.replace(character, "_")
    return title[:31] or "jsonl"


def convert_jsonl_to_excel(
    input_path: str | Path,
    output_path: str | Path,
    *,
    overwrite: bool = False,
) -> Path:
    """使用 pandas 将一个 JSONL 文件转换为可读的 Excel 文件。

    参数：
        input_path: JSONL 输入文件路径。
        output_path: 完整的 .xlsx 输出文件路径。
        overwrite: 是否允许覆盖已有输出文件，默认不覆盖。

    返回：实际生成的 Excel 文件路径。
    """

    try:
        import pandas as pd
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise RuntimeError("缺少 pandas 或 openpyxl 依赖，请先安装 requirements.txt") from exc

    source = Path(input_path).expanduser()
    target = Path(output_path).expanduser()
    if not source.is_file():
        raise FileNotFoundError(f"找不到输入文件：{source}")
    if source.suffix.lower() != ".jsonl":
        raise ValueError(f"输入文件必须使用 .jsonl 扩展名：{source}")
    if target.suffix.lower() != ".xlsx":
        raise ValueError(f"输出文件必须使用 .xlsx 扩展名：{target}")
    if target.exists() and not overwrite:
        raise FileExistsError(f"输出文件已存在：{target}")

    rows, columns = _read_jsonl(source)
    flattened_rows: list[dict[str, Any]] = []
    truncated_count = 0
    for row in rows:
        flattened: dict[str, Any] = {}
        for column in columns:
            value = _json_cell_value(row.get(column)) if column in row else None
            value, truncated = _truncate_cell_value(value)
            truncated_count += int(truncated)
            flattened[column] = value
        flattened_rows.append(flattened)

    target.parent.mkdir(parents=True, exist_ok=True)
    dataframe = pd.DataFrame.from_records(flattened_rows, columns=columns)
    sheet_name = _sheet_title(source)

    with pd.ExcelWriter(target, engine="openpyxl") as writer:
        dataframe.to_excel(writer, index=False, sheet_name=sheet_name, na_rep="")
        sheet = writer.book[sheet_name]
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

        header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for column_cells in sheet.iter_cols():
            longest = max((len(str(cell.value or "")) for cell in column_cells), default=0)
            width = min(60, max(10, longest + 2))
            sheet.column_dimensions[column_cells[0].column_letter].width = width
            for cell in column_cells[1:]:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    print(f"已转换 {len(rows)} 条记录、{len(columns)} 列。")
    if truncated_count:
        print(f"警告：有 {truncated_count} 个单元格超过 Excel 单元格长度限制，已截断。")
    print(f"Excel 文件：{target}")
    return target


def interactive_main() -> Path | None:
    """通过终端交互读取路径并执行一次转换。"""

    input_text = input("请输入 JSONL 文件路径：")
    output_text = input("请输入 Excel 输出路径（完整 .xlsx 文件路径）：")

    try:
        input_path = _clean_path(input_text)
        output_path = _clean_path(output_text)
        return convert_jsonl_to_excel(input_path, output_path)
    except FileExistsError as exc:
        answer = input(f"{exc}\n是否覆盖已有文件？请输入 y 继续，其他内容取消：").strip().lower()
        if answer not in {"y", "yes"}:
            print("已取消转换。")
            return None
        try:
            return convert_jsonl_to_excel(input_path, output_path, overwrite=True)
        except (FileNotFoundError, OSError, RuntimeError, ValueError) as retry_exc:
            print(f"转换失败：{retry_exc}")
            return None
    except (FileNotFoundError, OSError, RuntimeError, ValueError) as exc:
        print(f"转换失败：{exc}")
        return None


if __name__ == "__main__":
    interactive_main()

