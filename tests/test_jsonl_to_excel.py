"""交互式 JSONL 转 Excel 工具测试。"""

from __future__ import annotations

import json
import sys
import unittest
from io import StringIO
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch

from openpyxl import load_workbook

from jsonl_to_excel import convert_jsonl_to_excel, interactive_main


class JsonlToExcelTests(unittest.TestCase):
    def write_jsonl(self, path: Path, rows: list[object]) -> None:
        """把测试记录写成 UTF-8 JSONL 文件。"""
        path.write_text(
            "\n".join(json.dumps(row, ensure_ascii=False) for row in rows) + "\n",
            encoding="utf-8",
        )

    def test_converts_rows_with_union_columns_and_nested_values(self):
        """验证联合列和嵌套值能够正确写入 Excel。"""
        with TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "records.jsonl"
            output = root / "nested" / "records.xlsx"
            self.write_jsonl(
                source,
                [
                    {"id": 1, "name": "张三", "details": {"city": "北京"}, "tags": ["a", "b"]},
                    {"id": 2, "name": "李四", "extra": "第二行"},
                ],
            )

            result = convert_jsonl_to_excel(source, output)

            self.assertEqual(result, output)
            workbook = load_workbook(output)
            sheet = workbook.active
            self.assertEqual(list(sheet.values), [
                ("id", "name", "details", "tags", "extra"),
                (1, "张三", '{"city": "北京"}', '["a", "b"]', None),
                (2, "李四", None, None, "第二行"),
            ])
            self.assertEqual(sheet.freeze_panes, "A2")
            self.assertEqual(sheet.auto_filter.ref, "A1:E3")

    def test_skips_blank_lines_and_reports_invalid_json_line(self):
        """验证空行被跳过且非法 JSON 会报告准确行号。"""
        with TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "records.jsonl"
            source.write_text('{"id": 1}\n\nnot-json\n', encoding="utf-8")

            with self.assertRaisesRegex(ValueError, r"第 3 行.*JSON"):
                convert_jsonl_to_excel(source, root / "records.xlsx")

    def test_rejects_non_object_json_records(self):
        """验证非对象 JSON 记录会被明确拒绝。"""
        with TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "records.jsonl"
            self.write_jsonl(source, [["not", "an", "object"]])

            with self.assertRaisesRegex(ValueError, r"第 1 行.*JSON 对象"):
                convert_jsonl_to_excel(source, root / "records.xlsx")

    def test_truncates_oversized_text_and_emits_warning(self):
        """验证超长单元格文本被截断并输出警告。"""
        with TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "records.jsonl"
            output = root / "records.xlsx"
            self.write_jsonl(source, [{"text": "甲" * 40_000}])

            stdout = StringIO()
            with patch.object(sys, "stdout", stdout):
                convert_jsonl_to_excel(source, output)

            value = load_workbook(output)["records"]["A2"].value
            self.assertLessEqual(len(value), 32_767)
            self.assertIn("已截断", value)
            self.assertIn("超过 Excel 单元格长度限制", stdout.getvalue())

    def test_missing_input_and_existing_output_are_safe(self):
        """验证缺失输入和已存在输出均采用安全失败策略。"""
        with TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            missing = root / "missing.jsonl"
            with self.assertRaises(FileNotFoundError):
                convert_jsonl_to_excel(missing, root / "records.xlsx")

            source = root / "records.jsonl"
            output = root / "records.xlsx"
            self.write_jsonl(source, [{"id": 1}])
            convert_jsonl_to_excel(source, output)
            with self.assertRaises(FileExistsError):
                convert_jsonl_to_excel(source, output)
            convert_jsonl_to_excel(source, output, overwrite=True)

    def test_interactive_main_reads_quoted_paths(self):
        """验证交互入口可以读取带引号和空格的路径。"""
        with TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "input records.jsonl"
            output = root / "output records.xlsx"
            self.write_jsonl(source, [{"message": "你好"}])

            answers = iter([f'"{source}"', f'"{output}"'])
            with patch("builtins.input", side_effect=lambda _: next(answers)):
                result = interactive_main()

            self.assertEqual(result, output)
            self.assertTrue(output.is_file())

    def test_interactive_main_reports_empty_path_without_traceback(self):
        """验证空路径只给出友好提示而不泄露异常堆栈。"""
        stdout = StringIO()
        with patch("builtins.input", side_effect=["", "output.xlsx"]), patch.object(sys, "stdout", stdout):
            result = interactive_main()

        self.assertIsNone(result)
        self.assertIn("转换失败", stdout.getvalue())


if __name__ == "__main__":
    unittest.main()
