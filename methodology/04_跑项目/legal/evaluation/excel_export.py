"""法律评测结果导出模块。

项目位置：法律 Benchmark 的“跑项目”环节，作为法律评测入口的本地辅助模块。
输入：法律评测生成的 legal_evaluation_results.jsonl 或其他逐行 JSON 结果文件。
输出：指定位置的 Excel 工作簿；每一行 JSON 对象对应表格中的一行。
上下游：上游是 evaluation/run.py 写出的 JSONL，下游是人工筛选、复盘和汇报。
副作用：读取 JSONL、创建父目录并覆盖同名 xlsx；不调用模型，也不访问外部项目。
"""

import json
from pathlib import Path
from typing import Any


def _flatten(value: Any) -> Any:
    """用途：把嵌套 JSON 值转换为 Excel 单元格可以直接保存的内容。

    输入：value 可以是字典、列表、字符串、数字、布尔值或 None。
    输出：字典和列表变成保留中文的 JSON 字符串，其他值原样返回。
    运行前数据形态：JSONL 记录中的某个字段可能是嵌套对象或数组。
    运行后数据变化：复杂字段被压成一个可读单元格，不改变原始 JSONL。
    副作用：只在内存中转换，不写文件、不调用模型。
    异常或失败处理：不可序列化的嵌套值由 json.dumps 抛出异常。
    """

    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def export_jsonl(
    input_path: str | Path,
    output_path: str | Path | None = None,
    max_items: int | None = None,
) -> Path:
    """用途：把法律评测 JSONL 导出为带表头和基础样式的 Excel 文件。

    输入：input_path 是一行一个 JSON 对象的文件；output_path 可指定 xlsx 路径；
    max_items 可限制导出的记录数。
    输出：返回实际生成的 Excel 路径。
    运行前数据形态：每个非空行是一个结果字典，字典之间的字段可以不同。
    运行后数据变化：所有记录出现过的字段成为列，嵌套值通过 _flatten 写入单元格。
    副作用：读取输入文件、创建父目录并覆盖同名 xlsx；依赖 openpyxl；不调用模型。
    异常或失败处理：输入不存在、JSON 非法或缺少 openpyxl 时抛出明确异常。
    最小示例：两行分别只有 verdict 和 latency 字段时，表头会同时包含这两列，
    缺失字段的单元格留空。
    """

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise RuntimeError("缺少 openpyxl 依赖") from exc

    source = Path(input_path)
    rows: list[dict[str, Any]] = []
    lines = source.read_text(encoding="utf-8").splitlines()
    for line in lines:
        if not line.strip():
            continue
        rows.append(json.loads(line))
        if max_items is not None and max_items > 0 and len(rows) >= max_items:
            break

    if output_path is None:
        target = source.with_suffix(".xlsx")
    else:
        target = Path(output_path)
    target.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = source.stem[:31]

    columns: list[str] = []
    for row in rows:
        for key in row:
            if key not in columns:
                columns.append(key)

    if columns:
        sheet.append(columns)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")

        for row in rows:
            cells: list[Any] = []
            for column in columns:
                value = row.get(column, "")
                cells.append(_flatten(value))
            sheet.append(cells)

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

        for column_cells in sheet.columns:
            longest_cell_length = 0
            for cell in column_cells:
                cell_text = str(cell.value or "")
                cell_length = len(cell_text)
                if cell_length > longest_cell_length:
                    longest_cell_length = cell_length

            max_length = min(60, longest_cell_length + 2)
            column_letter = column_cells[0].column_letter
            sheet.column_dimensions[column_letter].width = max(10, max_length)

            for cell in column_cells[1:]:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    workbook.save(target)
    return target
