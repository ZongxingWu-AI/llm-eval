"""评测 JSONL 转 Excel 工具。

项目位置：仓库公共 tools 层，可服务 C-Eval、Pairwise Judge 和法律线。
输入：单个 JSONL 路径，或三条评测线 data/results 下自动发现的 JSONL。
输出：字段作为列、记录作为行的 xlsx 工作簿。
上下游：上游是各评测线产生的结构化结果，下游是人工阅读、筛选和报告制作。
副作用：读取 JSONL 并覆盖目标 Excel 文件；不调用任何模型。"""

import argparse
import json
import sys
from pathlib import Path
from typing import Iterable



def _flatten(value):
    """用途：把 JSON 字段值转换为适合 Excel 单元格保存的文本或标量。

    输入：value 可以是字典、列表、字符串、数字、布尔值或 None。
    输出：字典和列表返回未转义中文的 JSON 字符串；其他值原样返回。
    运行前数据形态：运行前单元格值可能是嵌套对象。
    运行后数据变化：运行后复杂值成为一个可读字符串，避免拆散 schema。
    副作用：只处理内存，不写文件、不调用模型。
    异常或失败处理：不可 JSON 序列化的嵌套值会由 json.dumps 抛出异常。"""

    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def export_jsonl(input_path: str | Path, output_path: str | Path | None = None,
                 max_items: int | None = None) -> Path:
    """用途：把一行一个 JSON 对象的评测结果导出为带表头和基础样式的 Excel。

    输入：input_path 是 JSONL；output_path 可选；max_items 限制导出行数。
    输出：返回生成的 xlsx Path。
    运行前数据形态：运行前每个非空行是字段可能不同的 JSON 对象。
    运行后数据变化：运行后所有出现过的键成为列，嵌套值经 _flatten 写入单元格。
    副作用：读取 JSONL，创建父目录并覆盖同名 xlsx；依赖 openpyxl，不调用模型。
    异常或失败处理：输入不存在、某行 JSON 非法或缺少 openpyxl 时抛出明确异常。
    最小示例：两行分别含 a 和 b 字段时，Excel 表头会同时包含 a、b。"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise RuntimeError("缺少 openpyxl 依赖") from exc
    source = Path(input_path)
    rows = []
    for line in source.read_text(encoding="utf-8").splitlines():
        if line.strip():
            rows.append(json.loads(line))
        if max_items is not None and max_items > 0 and len(rows) >= max_items:
            break
    target = Path(output_path) if output_path else source.with_suffix(".xlsx")
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = source.stem[:31]
    columns: list[str] = []
    for row in rows:
        for key in row:
            if key not in columns:
                columns.append(key)
    if columns:
        sheet.append(columns)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        for row in rows:
            cells = []
            for column in columns:
                value = row.get(column, "")
                cells.append(_flatten(value))
            sheet.append(cells)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column_cells in sheet.columns:
            longest_cell_length = 0
            for cell in column_cells:
                cell_text = str(cell.value or "")
                cell_length = len(cell_text)
                if cell_length > longest_cell_length:
                    longest_cell_length = cell_length
            max_length = min(60, longest_cell_length + 2)
            sheet.column_dimensions[column_cells[0].column_letter].width = max(10, max_length)
            for cell in column_cells[1:]:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    workbook.save(target)
    return target


def discover_jsonl() -> Iterable[Path]:
    """用途：扫描三条评测线 data 和 results 中可批量导出的 JSONL 文件。

    输入：无参数；从 PROJECT_ROOT/tracks 下固定三个模块开始扫描。
    输出：按评测线和目录顺序逐个产生 Path。
    运行前数据形态：运行前各评测线文件分散在自己的 data/results。
    运行后数据变化：运行后调用方可逐个交给 export_jsonl。
    副作用：只读取目录结构，不读取文件内容、不写文件、不调用模型。
    异常或失败处理：目录不存在时跳过；无 JSONL 时产生空迭代。"""

    for track in ("ceval", "pairwise_judge", "legal_benchmark"):
        root = PROJECT_ROOT / "tracks" / track
        for folder_name in ("data", "results"):
            folder = root / folder_name
            if folder.exists():
                yield from sorted(folder.rglob("*.jsonl"))


def main() -> None:
    """用途：提供 Excel 导出 CLI，支持单文件模式和扫描三条评测线的批量模式。

    输入：--input、--output、--max-items 来自 argparse；不传 input 时自动扫描。
    输出：单文件打印目标路径；批量模式打印每个映射和总数。
    运行前数据形态：运行前是 JSONL 路径或空参数。
    运行后数据变化：运行后在源文件旁或指定位置生成 Excel。
    副作用：调用 export_jsonl 创建并覆盖 xlsx 文件；不调用模型。
    异常或失败处理：只传 --output 而不传 --input 时由 argparse 报错；单文件导出异常向上抛出。"""

    parser = argparse.ArgumentParser(description="把评测 JSONL 导出为 Excel")
    parser.add_argument("--input", help="单个 JSONL；不传则扫描三条评测线")
    parser.add_argument("--output", help="单文件输出 XLSX；批量模式不可用")
    parser.add_argument("--max-items", type=int, default=None, help="每个文件最多导出前 N 行")
    args = parser.parse_args()
    if args.input:
        target = export_jsonl(args.input, args.output, args.max_items)
        print(f"完成：{target}")
        return
    if args.output:
        parser.error("--output 只能与 --input 一起使用")
    count = 0
    for source in discover_jsonl():
        target = export_jsonl(source, max_items=args.max_items)
        print(f"{source} -> {target}")
        count += 1
    print(f"共导出 {count} 个文件")


if __name__ == "__main__":
    main()
